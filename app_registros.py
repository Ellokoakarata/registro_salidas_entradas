import streamlit as st
import pandas as pd
from datetime import datetime, timedelta
import pytz
import io
import openpyxl
from openpyxl.utils import get_column_letter
from openpyxl.styles import Border, Side

# NUEVOS IMPORTS ---------------------------------------------
import zipfile
import re
# ------------------------------------------------------------

# ---------------------------
# INICIALIZACIÓN DE FIREBASE
# ---------------------------
import firebase_admin
from firebase_admin import credentials, firestore, storage

firebase_secrets = st.secrets["firebase"]

if not firebase_admin._apps:
    cred = credentials.Certificate({
        "type": firebase_secrets["type"],
        "project_id": firebase_secrets["project_id"],
        "private_key_id": firebase_secrets["private_key_id"],
        "private_key": firebase_secrets["private_key"],
        "client_email": firebase_secrets["client_email"],
        "client_id": firebase_secrets["client_id"],
        "auth_uri": firebase_secrets["auth_uri"],
        "token_uri": firebase_secrets["token_uri"],
        "auth_provider_x509_cert_url": firebase_secrets["auth_provider_x509_cert_url"],
        "client_x509_cert_url": firebase_secrets["client_x509_cert_url"]
    })
    firebase_admin.initialize_app(cred, {
        'storageBucket': firebase_secrets["storageBucket"]
    })

db = firestore.client()
bucket = storage.bucket()

# ---------------------------
# CONFIGURACIÓN DE HORARIOS
# ---------------------------
ENTRY_DEADLINE = 11    # Se permite marcar entrada solo hasta las 11:00 AM
EXIT_START = 18        # Se permite marcar salida solo hasta las 6:00 PM

# ---------------------------
# FUNCIONES AUXILIARES EXISTENTES
# ---------------------------
def get_week_filename():
    """Genera el nombre del archivo según el año y la semana actual."""
    now = datetime.now()
    year, week, _ = now.isocalendar()
    return f"registro_{year}_W{week}.xlsx"

def utc_to_lima(utc_dt):
    """Convierte un datetime en UTC a la hora de Lima (America/Lima)."""
    lima_tz = pytz.timezone("America/Lima")
    return utc_dt.astimezone(lima_tz)

def format_datetime(dt):
    """Formatea el datetime a cadena, usando la hora de Lima."""
    return utc_to_lima(dt).strftime("%d/%m/%Y %I:%M:%S %p")

def parse_timedelta(td_str):
    """Convierte una cadena tipo 'H:MM:SS' a un objeto timedelta."""
    try:
        h, m, s = map(int, td_str.split(':'))
        return timedelta(hours=h, minutes=m, seconds=s)
    except Exception:
        return timedelta()

def create_summary_df(df):
    """Crea un DataFrame resumen con el total de horas trabajadas por cada trabajador."""
    summary = {}
    for name in df["Nombre"].unique():
        valid = df[(df["Nombre"] == name) & (df["Horas Trabajadas"].notna())]
        total = timedelta()
        for _, row in valid.iterrows():
            total += parse_timedelta(row["Horas Trabajadas"])
        summary[name] = str(total)
    summary_df = pd.DataFrame(list(summary.items()), columns=["Nombre", "Total Horas Trabajadas"])
    summary_df = summary_df.sort_values("Nombre")
    return summary_df

def save_week_data_and_upload(df, filename):
    """
    Guarda el DataFrame en un archivo Excel con dos hojas: 'Registros' y 'Resumen',
    ajusta las columnas, añade bordes a las celdas y sube el archivo directamente a Firebase Storage.
    Se utiliza un buffer en memoria, sin escribir en disco.
    """
    output = io.BytesIO()
    with pd.ExcelWriter(output, engine='openpyxl') as writer:
        df.to_excel(writer, sheet_name='Registros', index=False)
        ws = writer.sheets['Registros']
        # Autoajuste de columnas para 'Registros'
        for col_cells in ws.columns:
            max_length = 0
            col_letter = col_cells[0].column_letter
            for cell in col_cells:
                if cell.value:
                    cell_length = len(str(cell.value))
                    if cell_length > max_length:
                        max_length = cell_length
            ws.column_dimensions[col_letter].width = max_length + 2

        summary_df = create_summary_df(df)
        summary_df.to_excel(writer, sheet_name='Resumen', index=False)
        ws_summary = writer.sheets['Resumen']
        # Autoajuste de columnas para 'Resumen'
        for col_cells in ws_summary.columns:
            max_length = 0
            col_letter = col_cells[0].column_letter
            for cell in col_cells:
                if cell.value:
                    cell_length = len(str(cell.value))
                    if cell_length > max_length:
                        max_length = cell_length
            ws_summary.column_dimensions[col_letter].width = max_length + 2

        # Agregar bordes a todas las celdas en ambas hojas
        thin_border = Border(
            left=Side(style="thin"), 
            right=Side(style="thin"), 
            top=Side(style="thin"), 
            bottom=Side(style="thin")
        )
        for ws_sheet in [writer.sheets['Registros'], writer.sheets['Resumen']]:
            for row in ws_sheet.iter_rows():
                for cell in row:
                    cell.border = thin_border

    output.seek(0)
    blob = bucket.blob(filename)
    blob.upload_from_string(output.read(), content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    st.write(f"Archivo {filename} subido a Firebase Storage.")

def load_week_data(filename):
    """
    Intenta cargar el DataFrame de la hoja 'Registros' desde Firebase Storage.
    Si no existe, se crea uno nuevo y se sube.
    """
    blob = bucket.blob(filename)
    if blob.exists():
        data = blob.download_as_bytes()
        df = pd.read_excel(io.BytesIO(data), sheet_name='Registros')
        return df
    else:
        df = pd.DataFrame(columns=["Nombre", "Fecha", "Entrada", "Salida", "Horas Trabajadas"])
        save_week_data_and_upload(df, filename)
        st.write(f"Nuevo archivo para la semana creado: {filename}")
        return df

def update_firestore(worker, data):
    """
    Actualiza Firestore en la colección 'registros'.
    Cada trabajador tendrá un documento cuyo ID es su nombre, y se guardan los registros.
    """
    doc_ref = db.collection("registros").document(worker)
    doc_ref.set(data)

def register_event(worker, event_type):
    """
    Registra una entrada o salida para un trabajador.
    Se actualiza Firestore y se actualiza el archivo Excel en Firebase Storage.
    Se convierte la hora de UTC a la hora de Lima.
    Se validan horarios:
      - Entrada: solo se permite hasta las 11:00 AM.
      - Salida: solo se permite hasta las 6:00 PM.
    Si no se marcó entrada, no se permite marcar salida.
    """
    filename = get_week_filename()
    df = load_week_data(filename)
    today_str = datetime.now().strftime("%Y-%m-%d")
    
    now_utc = datetime.now(pytz.utc)
    local_now = utc_to_lima(now_utc)
    now_str = format_datetime(now_utc)
    
    # Validar horario según tipo de evento
    if event_type == "entrada":
        if local_now.hour >= ENTRY_DEADLINE:
            return False, "Fuera del horario permitido para marcar entrada (hasta las 11:00 AM)."
    elif event_type == "salida":
        if local_now.hour > EXIT_START or (local_now.hour == EXIT_START and local_now.minute > 0):
            return False, "Fuera del horario permitido para marcar salida (hasta las 6:00 PM)."
    
    record = df[(df["Nombre"] == worker) & (df["Fecha"] == today_str)]
    
    if event_type == "entrada":
        if not record.empty and pd.notna(record.iloc[0]["Entrada"]):
            return False, "Ya se ha registrado una entrada hoy para este trabajador."
        if record.empty:
            new_row = {
                "Nombre": worker,
                "Fecha": today_str,
                "Entrada": now_str,
                "Salida": "No marcó salida",
                "Horas Trabajadas": "No marcó salida"
            }
            df = pd.concat([df, pd.DataFrame([new_row])], ignore_index=True)
        else:
            idx = record.index[0]
            df.at[idx, "Entrada"] = now_str
        save_week_data_and_upload(df, filename)
        update_firestore(worker, {"Fecha": today_str, "Evento": "entrada", "Timestamp": now_str})
        return True, f"Entrada registrada para {worker} a las {now_str}"
    
    elif event_type == "salida":
        if record.empty or pd.isna(record.iloc[0]["Entrada"]):
            return False, "No se ha registrado entrada hoy para este trabajador."
        if pd.notna(record.iloc[0]["Salida"]) and record.iloc[0]["Salida"] != "No marcó salida":
            return False, "Ya se ha registrado una salida hoy para este trabajador."
        
        idx = record.index[0]
        df.at[idx, "Salida"] = now_str
        try:
            entry_time = datetime.strptime(df.at[idx, "Entrada"], "%d/%m/%Y %I:%M:%S %p")
            exit_time = datetime.strptime(now_str, "%d/%m/%Y %I:%M:%S %p")
            worked = exit_time - entry_time
            df.at[idx, "Horas Trabajadas"] = str(worked)
        except Exception as e:
            return False, f"Error al calcular las horas trabajadas: {e}"
        save_week_data_and_upload(df, filename)
        update_firestore(worker, {"Fecha": today_str, "Evento": "salida", "Timestamp": now_str})
        return True, f"Salida registrada para {worker} a las {now_str}"
    
    return False, "Evento desconocido."

def get_worker_week_hours(worker):
    """Obtiene la suma de las horas trabajadas en la semana para un trabajador."""
    filename = get_week_filename()
    df = load_week_data(filename)
    records = df[(df["Nombre"] == worker) & (df["Horas Trabajadas"].notna())]
    total = timedelta()
    for _, row in records.iterrows():
        total += parse_timedelta(row["Horas Trabajadas"])
    return total

# ---------------------------
# NUEVA FUNCIÓN: GENERAR ARCHIVO MENSUAL
# ---------------------------
def generate_monthly_file(selected_year, selected_month):
    """
    Reúne todos los registros de los archivos semanales almacenados en Firebase Storage correspondientes
    al mes y año seleccionados. Filtra los registros en base a la columna "Fecha" (formato YYYY-MM-DD).
    Genera un archivo Excel con dos hojas: "Registros" y "Resumen", manteniendo el estilo y formato.
    Retorna el contenido binario del archivo.
    """
    blobs = list(bucket.list_blobs(prefix="registro_"))  # lista de blobs
    monthly_dfs = []
    for blob in blobs:
        try:
            data = blob.download_as_bytes()
            df_week = pd.read_excel(io.BytesIO(data), sheet_name='Registros')
            if not df_week.empty:
                df_week["Fecha_dt"] = pd.to_datetime(df_week["Fecha"], format="%Y-%m-%d", errors="coerce")
                mask = (df_week["Fecha_dt"].dt.year == selected_year) & (df_week["Fecha_dt"].dt.month == selected_month)
                df_filtered = df_week.loc[mask].drop(columns=["Fecha_dt"])
                if not df_filtered.empty:
                    monthly_dfs.append(df_filtered)
        except Exception as e:
            st.error(f"Error procesando el archivo {blob.name}: {e}")
    
    if not monthly_dfs:
        st.error("No se encontraron registros para el mes y año seleccionados.")
        return None

    df_month = pd.concat(monthly_dfs, ignore_index=True)
    resumen_month = create_summary_df(df_month)
    
    output = io.BytesIO()
    with pd.ExcelWriter(output, engine='openpyxl') as writer:
        df_month.to_excel(writer, sheet_name='Registros', index=False)
        ws = writer.sheets['Registros']
        for col_cells in ws.columns:
            max_length = 0
            col_letter = col_cells[0].column_letter
            for cell in col_cells:
                if cell.value:
                    cell_length = len(str(cell.value))
                    if cell_length > max_length:
                        max_length = cell_length
            ws.column_dimensions[col_letter].width = max_length + 2

        resumen_month.to_excel(writer, sheet_name='Resumen', index=False)
        ws_resumen = writer.sheets['Resumen']
        for col_cells in ws_resumen.columns:
            max_length = 0
            col_letter = col_cells[0].column_letter
            for cell in col_cells:
                if cell.value:
                    cell_length = len(str(cell.value))
                    if cell_length > max_length:
                        max_length = cell_length
            ws_resumen.column_dimensions[col_letter].width = max_length + 2

        thin_border = Border(
            left=Side(style="thin"),
            right=Side(style="thin"),
            top=Side(style="thin"),
            bottom=Side(style="thin")
        )
        for ws_sheet in [ws, ws_resumen]:
            for row in ws_sheet.iter_rows():
                for cell in row:
                    cell.border = thin_border

    output.seek(0)
    return output

# ---------------------------
# NUEVAS FUNCIONES AUXILIARES PARA DESCARGAS ----------------
_pattern_week = re.compile(r"registro_(\d{4})_W(\d{1,2})\.xlsx")

def list_week_files() -> list[str]:
    """Devuelve los blobs con patrón registro_YYYY_Www.xlsx (solo semanas)."""
    return sorted(
        [b.name for b in bucket.list_blobs(prefix="registro_") if _pattern_week.match(b.name)]
    )

def week_files_for_month(year: int, month: int, all_files: list[str]) -> list[str]:
    """Filtra las semanas cuyo primer día ISO-week cae en el mes/año dados."""
    target = []
    for fname in all_files:
        year_w, week = map(int, _pattern_week.match(fname).groups())
        first_day = datetime.fromisocalendar(year_w, week, 1)
        if first_day.year == year and first_day.month == month:
            target.append(fname)
    return target

def zip_blobs(blob_names: list[str]) -> io.BytesIO:
    """Crea un ZIP en memoria con los blobs dados y lo devuelve."""
    buf = io.BytesIO()
    with zipfile.ZipFile(buf, "w", zipfile.ZIP_DEFLATED) as zf:
        for name in blob_names:
            data = bucket.blob(name).download_as_bytes()
            zf.writestr(name, data)
    buf.seek(0)
    return buf

# ---------------------------
# INTERFAZ DE USUARIO CON STREAMLIT
# ---------------------------

user_list = ["Nelida Ruiz", "Ricardo Adrian Ruiz", "Paula Lecaros"]
user_passwords = st.secrets["user_passwords"]

st.title("Registro de Entradas y Salidas (CLOUD: Firestore + Firebase Storage)")

with st.expander("Selecciona al Trabajador"):
    worker = st.selectbox("Elige tu nombre:", [""] + user_list)

if worker:
    password_input = st.text_input("Ingrese su contraseña:", type="password")
    
    if password_input:
        if password_input == user_passwords.get(worker, ""):
            if worker == "Ricardo Adrian Ruiz":
                st.info("Bienvenido, ADMIN.")
            else:
                st.info(f"Bienvenido, {worker}.")
            
            st.header(f"Registro para: {worker}")
            
            col1, col2 = st.columns(2)
            with col1:
                if st.button("Registrar Entrada"):
                    success, msg = register_event(worker, "entrada")
                    if success:
                        st.success(msg)
                    else:
                        st.warning(msg)
            with col2:
                if st.button("Registrar Salida"):
                    success, msg = register_event(worker, "salida")
                    if success:
                        st.success(msg)
                    else:
                        st.warning(msg)
            
            total_hours = get_worker_week_hours(worker)
            st.write("Total de horas trabajadas esta semana:", str(total_hours))
            
            if worker == "Ricardo Adrian Ruiz":
                st.subheader("Resumen Semanal General")
                if st.button("Mostrar resumen de horas por trabajador"):
                    filename = get_week_filename()
                    df = load_week_data(filename)
                    resumen = create_summary_df(df)
                    st.dataframe(resumen)
            else:
                if st.button("Mostrar mis registros semanales"):
                    filename = get_week_filename()
                    df = load_week_data(filename)
                    worker_records = df[df["Nombre"] == worker]
                    st.dataframe(worker_records)
            
            # --- Sección ADMIN: Descarga de registros semanales ---------------
            if worker == "Ricardo Adrian Ruiz":
                st.markdown("---")
                st.subheader("Descarga de registros semanales")

                week_files = list_week_files()

                if not week_files:
                    st.info("No hay archivos semanales en Firebase Storage.")
                else:
                    selected_file = st.selectbox(
                        "Selecciona un archivo semanal para descargar:", [""] + week_files
                    )

                    col_dl_one, col_dl_month, col_dl_all = st.columns(3)

                    # Descargar archivo individual
                    with col_dl_one:
                        if selected_file:
                            data = bucket.blob(selected_file).download_as_bytes()
                            st.download_button(
                                label=f"Descargar {selected_file}",
                                data=data,
                                file_name=selected_file,
                                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
                            )

                    # ZIP del mes seleccionado (usa año/mes elegidos arriba)
                    with col_dl_month:
                        if st.button("ZIP del mes seleccionado"):
                            month_weeks = week_files_for_month(
                                selected_year, selected_month, week_files
                            )
                            if month_weeks:
                                buf = zip_blobs(month_weeks)
                                zip_name = f"registros_{selected_year}_{selected_month:02d}.zip"
                                st.download_button(
                                    label=f"Descargar {zip_name}",
                                    data=buf,
                                    file_name=zip_name,
                                    mime="application/zip"
                                )
                            else:
                                st.warning("No hay semanas para ese mes.")

                    # ZIP con todo el histórico
                    with col_dl_all:
                        if st.button("ZIP con TODO"):
                            buf = zip_blobs(week_files)
                            st.download_button(
                                label="Descargar registros_all.zip",
                                data=buf,
                                file_name="registros_all.zip",
                                mime="application/zip"
                            )
            
            # --- Sección ADMIN: Generar y descargar archivo mensual -----------
            if worker == "Ricardo Adrian Ruiz":
                st.markdown("---")
                st.subheader("Archivo Mensual (nuevo, generado al vuelo)")

                col_year, col_month = st.columns(2)
                with col_year:
                    current_year = datetime.now().year
                    selected_year = st.number_input(
                        "Elige el año", min_value=2000, max_value=current_year+1,
                        value=current_year, step=1
                    )
                with col_month:
                    month_names = [
                        "Enero", "Febrero", "Marzo", "Abril", "Mayo", "Junio",
                        "Julio", "Agosto", "Septiembre", "Octubre", "Noviembre", "Diciembre"
                    ]
                    selected_month_name = st.selectbox("Elige el mes", month_names)
                    selected_month = month_names.index(selected_month_name) + 1
                
                if st.button("Generar archivo mensual"):
                    output_buffer = generate_monthly_file(selected_year, selected_month)
                    if output_buffer is not None:
                        filename = f"registro_{selected_year}_{selected_month:02d}.xlsx"
                        st.download_button(
                            label="Descargar archivo mensual",
                            data=output_buffer,
                            file_name=filename,
                            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
                        )
        else:
            st.error("Contraseña incorrecta. Intente nuevamente.")
